import sqlite3
from tkinter import ttk
from tkinter import *
from tkinter.ttk import *
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
import reporte_bd
from tkinter import messagebox
from tkcalendar import DateEntry
from datetime import datetime, date
import openpyxl
from openpyxl.styles import Font
from tkinter import filedialog

resultados_tree = []
lista_desplegables = []


def ventanaReporte(raiz):

    global resultados_tree

    vent_Reporte = Toplevel(raiz)
    vent_Reporte.title("Generar reporte de equipo")

    etiqueta_Reporte = Label(vent_Reporte, text="Para generar el reporte elige el principal criterio de búsqueda.")
    etiqueta_Reporte.grid(row=0, column=0, pady=10)

    dict_reporte = {"": "TODOS", "MARCA": "MARCA", "MODELO": "MODELO", "CONFIGURACION": "CONF. GASES", "FECHA_CAL": "FECHA CAL",
                    "SIG_FECHA_CAL": "PROX. FECHA CAL", "EMPRESA": "RAZON SOCIAL", "RUC": "RUC", "ZONA": "PROVINCIA", "ESTADO": "ESTADO EQUIPOS",
                    "COTIZACION": "NRO. COTIZACIÓN", "ORDEN": "POR OC", "FACTURA": "FACTURADOS"}

    dict_reporte = {value: key for key, value in dict_reporte.items()}  # Invertir los valores de key:value
    opciones_reporte = list(dict_reporte.keys())

    frame_Lista_desplegable = Frame(vent_Reporte)
    frame_Lista_desplegable.grid(row=1, column=0, columnspan=2)

    lista_desplegable_variable = StringVar()

    # La lista desplegable para los valores del diccionario "dict_reporte"
    lista_desplegable = Combobox(frame_Lista_desplegable, values=opciones_reporte, state="readonly", textvariable=lista_desplegable_variable)
    lista_desplegable.set(opciones_reporte[6])
    lista_desplegable.grid(row=0, column=0, columnspan=2, pady=10, padx=2)

    #Inicializa la lista de criterios disponibles en envia el argumento del "Key" del diccionario.
    valores_lista_criterio_act = actualizar_lista_criterio(dict_reporte.get(lista_desplegable_variable.get()),frame_Lista_desplegable)


    etiqueta_Entrada = Label(frame_Lista_desplegable, text="Filtrar por: ")
    etiqueta_Entrada.grid(row=1, column=0, padx=2, pady=2, sticky="E")


    cuadro_Entrada = Combobox(frame_Lista_desplegable, values=valores_lista_criterio_act, width=40)                     # Valor de RUC de entrada para buscar en BD
    cuadro_Entrada.grid(row=1, column=1, padx=2, pady=2, sticky="W")

    # Actualiza los valores de la lista deplegable.
    lista_desplegable.bind("<<ComboboxSelected>>", lambda event: cuadro_Entrada.config(values=actualizar_lista_criterio(dict_reporte.get(lista_desplegable_variable.get()), frame_Lista_desplegable)))

    etiqueta_encabezados = Label(frame_Lista_desplegable, text="Elige Nro. criterios para el reporte")
    etiqueta_encabezados.grid(row=4, column=0, pady=10, padx=5)

    Cantidad_columnas = ["1", "2", "3", "4", "5", "6", "7", "8"]
    lista_columnas = Combobox(frame_Lista_desplegable, values=Cantidad_columnas, state="readonly")
    lista_columnas.set(Cantidad_columnas[4])
    lista_columnas.grid(row=4, column=1, pady=10, padx=2)

    columna_encabezado = StringVar()
    columna_encabezado = ["SERIE", "MARCA", "MODELO", "CONFIGURACION", "FECHA_CAL", "SIG_FECHA_CAL", "RUC", "EMPRESA", "ZONA", "ESTADO", "COTIZACION", "ORDEN", "FACTURA"]
    ventana_criterios_busqueda = Frame(vent_Reporte)
    ventana_criterios_busqueda.grid(row=2, column=0)

    lista_columnas.bind("<<ComboboxSelected>>", lambda event: criterios_busqueda(lista_columnas.get(), ventana_criterios_busqueda, columna_encabezado))

    boton_generar_reporte = Button(frame_Lista_desplegable, text="Generar Reporte", command=lambda: generarReporte(lista_desplegable_variable.get(), cuadro_Entrada.get(), resultados_tree, lista_desplegables, dict_reporte))
    boton_generar_reporte.grid(row=5, column=0, columnspan=2, pady=10, padx=2)

    Marco_botones=Frame(vent_Reporte)
    Marco_botones.grid(row=3, column=0)

    boton_exportar_pdf = Button(Marco_botones, text="Exportar a PDF", command=lambda: exportarPDF(lista_desplegable_variable.get(), cuadro_Entrada.get(), resultados_tree, lista_desplegables))
    boton_exportar_pdf.grid(row=0, column=1, pady=10, padx=15)

    boton_exportar_excel = Button(Marco_botones, text="Exportar a Excel", command=lambda: exportarExcel(lista_desplegable_variable.get(), cuadro_Entrada.get(), resultados_tree, lista_desplegables))
    boton_exportar_excel.grid(row=0, column=0, pady=10, padx=15)

def actualizar_lista_criterio(valor_criterio, frame):

    miConexion = sqlite3.connect("Equipos")
    miCursor = miConexion.cursor()
    global fecha_1, fecha_2

    if valor_criterio == "FECHA_CAL" or valor_criterio == "SIG_FECHA_CAL":
        fecha_1 = DateEntry(frame, width=17, background='darkblue', foreground='white', date_pattern='dd/mm/yyyy')
        fecha_1.grid(row=2, column=1, padx=5, pady=5)
        fecha_1.config(justify="center")
        fecha_1_nombre = Label(frame, text="Fecha inicio:")
        fecha_1_nombre.grid(row=2, column=0, sticky="e", padx=10, pady=10)

        fecha_2 = DateEntry(frame, width=17, background='darkblue', foreground='white', date_pattern='dd/mm/yyyy')
        fecha_2.grid(row=3, column=1, padx=5, pady=5)
        fecha_2.config(justify="center")
        fecha_2_nombre = Label(frame, text="Fecha Fin:")
        fecha_2_nombre.grid(row=3, column=0, sticky="e", padx=10, pady=10)

    else:
        try:
            consulta = f"SELECT {valor_criterio} FROM DATOS_EQUIPOS"
            miCursor.execute(consulta)

            resultados_lista_criterio = miCursor.fetchall()

            valores_lista_criterio = sorted(set(row[0] for row in resultados_lista_criterio if row[0]))
            return valores_lista_criterio
        except sqlite3.OperationalError:
            pass
    miConexion.close()

def criterios_busqueda(valor, frame, encabezado):

    global resultados_tree

    for aplicaciones in frame.winfo_children():
        aplicaciones.destroy()

    for val in range(int(valor)):
        lista_desplegable2 = Combobox(frame, values=encabezado, state="readonly", width=16)
        lista_desplegable2.set(encabezado[0])
        lista_desplegable2.grid(row=0, column=val, pady=10, padx=2)

     # vista de arbol para visualizar los datos extraidos como tabla.

    for listas in frame.winfo_children():
        listas.bind("<<ComboboxSelected>>", lambda event: actualizar_encabezados(frame, valor))


def actualizar_encabezados(frame, valor):

    global lista_desplegables
    global resultados_tree

    lista_desplegables = []

    for lista in frame.winfo_children():
        try:
            lista_desplegables.append(lista.get())
        except AttributeError:
            pass

    resultados_tree = Treeview(frame, columns=lista_desplegables, show="headings")
    for i in range(len(lista_desplegables)):
        resultados_tree.heading(i, text=lista_desplegables[i])
        resultados_tree.column(i, width=120)
    resultados_tree.grid(row=1, column=0, columnspan=int(valor), padx=5, pady=10)


def generarReporte(criterio_principal, criterio_busqueda, resultado_tree, string_consulta, dict):

    valor = dict.get(criterio_principal)

    # Aquí puedes implementar la lógica para generar el reporte dinamico según el criterio seleccionado
    print(f"Generando reporte: {criterio_principal}")
    txt_consulta = ", ".join(string_consulta)
    miConexion = sqlite3.connect("Equipos")
    miCursor = miConexion.cursor()

    try:
        if criterio_principal == "TODOS":
            query = f"SELECT {txt_consulta} FROM DATOS_EQUIPOS ORDER BY EMPRESA"
            miCursor.execute(query)
            resultados = miCursor.fetchall()

        elif criterio_principal == "FECHA CAL" or criterio_principal == "PROX. FECHA CAL":

            fecha_inicio = fecha_1.get_date()
            fecha_fin = fecha_2.get_date()

            query2 = f"SELECT {txt_consulta} FROM DATOS_EQUIPOS WHERE {valor} BETWEEN ? AND ? order by {valor}"
            parametros = (fecha_inicio, fecha_fin,)

            print("Consulta SQL:", query2, "Parametros:", parametros)

            miCursor.execute(query2, parametros)
            resultados = miCursor.fetchall()

        else:
            # Consulta para obtener equipos por el criterio de busqueda
            query3 = f"SELECT {txt_consulta} FROM DATOS_EQUIPOS WHERE {valor} = ?"

            parametros = (criterio_busqueda,)   # La coma al final es una convecion para garantizar que se cree una tupla.

            miCursor.execute(query3, parametros)
            resultados = miCursor.fetchall()

        # Limpiar la tabla.
        for columna in resultado_tree.get_children():
            resultado_tree.delete(columna)

        # Mostrar los resultados
        for resultado in resultados:
            resultado_tree.insert("", "end", values=resultado)

    except sqlite3.Error as error:
        print(f"Error en la consulta SQL: {error}")

    finally:
        miConexion.close()

def exportarPDF(criterio1, criterio2, treeview, header):
    try:
        resultados = treeview.get_children()
        if not resultados:
            messagebox.showinfo("Generar PDF", "No hay datos para exportar")

        # Encabezados
        lista_completa = []

        for resultado in resultados:
            valores = treeview.item(resultado, "values")
            valores_list = list(valores)  # Convierte la tupla a lista
            lista_completa.append(valores_list)

        ahora = datetime.now()
        ahora_string = ahora.strftime("%d-%m-%Y %H-%M-%S")

        nombre_reporte = "reportes\Reporte_%s_%s.pdf" % (criterio1, ahora_string)
        reporte_bd.PDFPSReporte(nombre_reporte, lista_completa, header, criterio1, criterio2)
        messagebox.showinfo("Generar PDF", "Reporte creado")
    except AttributeError:
        messagebox.showinfo("Generar PDF", "No hay datos para exportar")


def exportarExcel(criterio1, treeview, header):
    try:
        resultados = treeview.get_children()
        if not resultados:
            messagebox.showinfo("Generar Excel", "No hay datos para exportar")

        # Crear un nuevo libro de trabajo y una hoja
        libro = openpyxl.Workbook()
        hoja = libro.active

        # Encabezados
        for col, valor in enumerate(header, 1):
            hoja.cell(row=1, column=col, value=valor)
            hoja.cell(row=1, column=col).font = Font(bold=True)

        # Datos
        for row, resultado in enumerate(resultados, 2):
            valores = treeview.item(resultado, "values")
            for col, valor in enumerate(valores, 1):
                hoja.cell(row=row, column=col, value=valor)

        ahora = datetime.now()
        ahora_string = ahora.strftime("%d-%m-%Y %H-%M-%S")

        # Guardar el archivo
        nombre_archivo = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                       filetypes=[("Archivos Excel", "*.xlsx")],
                                                       initialfile=f"Reporte_{criterio1}_{ahora_string}")

        if nombre_archivo:  # Verificar si se seleccionó un archivo
            libro.save(nombre_archivo)
            messagebox.showinfo("Generar Excel", "Datos exportados a Excel correctamente")

    except AttributeError:
        messagebox.showinfo("Generar Excel", "No hay datos para exportar")




